import os

os.environ['TK_SILENCE_DEPRECATION'] = '1'

import tkinter as tk
import openpyxl
import tkinter as tk
from datetime import date

# Load or create the Excel file and sheet as before
current_directory = os.getcwd()

# Specify the relative path to the Excel file inside the 'Vacancy Tracker' folder
file_path = os.path.join(current_directory, 'Vacancy Tracker', 'vacancy_tracker.xlsx')

# Load or create the Excel file and sheet
try:
    workbook = openpyxl.load_workbook(file_path)
except FileNotFoundError:
    workbook = openpyxl.Workbook()

if 'Vacancies' in workbook.sheetnames:
    sheet = workbook['Vacancies']
else:
    sheet = workbook.active
    sheet.title = 'Vacancies'
    sheet['A1'] = 'Agency'
    sheet['B1'] = 'Job Title'
    sheet['C1'] = 'Date Posted'
    sheet['D1'] = 'Date Filled'

# Function to add a vacancy from the form
def add_vacancy():
    agency = agency_entry.get()
    job_title = job_title_entry.get()
    row = [agency, job_title, date.today(), None]
    sheet.append(row)
    try:
        workbook.save(file_path)
        print("Data added and file saved successfully!")
    except Exception as e:
        print(f"Error: {e}")
    agency_entry.delete(0, tk.END)
    job_title_entry.delete(0, tk.END)

# Create a Tkinter window and form
root = tk.Tk()
root.title('Vacancy Tracker Form')

agency_label = tk.Label(root, text='Agency')
agency_label.grid(row=0, column=0)
agency_entry = tk.Entry(root)
agency_entry.grid(row=0, column=1)

job_title_label = tk.Label(root, text='Job Title')
job_title_label.grid(row=1, column=0)
job_title_entry = tk.Entry(root)
job_title_entry.grid(row=1, column=1)

add_button = tk.Button(root, text='Add Vacancy', command=add_vacancy)
add_button.grid(row=2, column=0, columnspan=2)

root.mainloop()